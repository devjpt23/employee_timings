from openpyxl import load_workbook
from openpyxl.styles import PatternFill,NamedStyle

book = load_workbook('EmpData.xlsx')
sheet = book.active

column = sheet['B']
columnNumbers = 0

for i in column:
    columnNumbers += 1

for i in range(2, columnNumbers + 1):
    sheet[f'D{i}'] = f"=C{i}-B{i}"

redFill = PatternFill(patternType="solid", fgColor='C64747')
time_style = NamedStyle(name='time_style', number_format='[h]:mm')

sheet['H3'] = "=SUM(D2:D6)"
sheet['H3'].style = time_style


for row in range(2, 17):
    cell = sheet[f'C{row}']
    if cell.value is None:
        cell.fill = redFill
        sheet[f"D{row}"] = '0'
        
book.save('newEmpData.xlsx')
print("Saved changes!!!")
