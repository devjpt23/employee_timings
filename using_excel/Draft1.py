import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

book = load_workbook('EmpData.xlsx')
sheet = book.active

column = sheet['B']
columnNumbers = 0

for i in column:
    columnNumbers += 1

for i in range(2, columnNumbers + 1):
    sheet[f'D{i}'] = f"=C{i}-B{i}"
  

book.save('newEmp.xlsx')
print('changes have been made!')
